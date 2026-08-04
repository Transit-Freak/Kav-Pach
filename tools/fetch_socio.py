#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""שליפת המדד החברתי-כלכלי של הלמ"ס (אשכול 1–10 לכל יישוב) ממאגר
המידע הממשלתי (data.gov.il) — רץ על runner של GitHub (לסביבה המקומית
אין גישת רשת לדומיין).

פלט: parks/data/socio.json — {"year":…, "by_city":{שם יישוב: {"c":אשכול,
"r":דירוג}}, "source":…}
"""
import json
import re
import sys
import urllib.parse
import urllib.request

API = 'https://data.gov.il/api/3/action/'


def call(action, **params):
    import time
    url = API + action + '?' + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={'User-Agent': 'kav-bochan-data/1.0'})
    last = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=90) as r:
                d = json.load(r)
            if not d.get('success'):
                raise RuntimeError(f'{action} נכשל')
            return d['result']
        except Exception as e:
            last = e
            time.sleep(3 * (attempt + 1))
    raise last


def find_field(fields, *words):
    for f in fields:
        fid = f.get('id', '')
        if all(w in fid for w in words):
            return fid
    return None


def fetch_bytes(url):
    # UA דפדפני — אתר הלמ"ס חוסם לקוחות אוטומטיים אנונימיים
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def parse_tabular(content, fmt):
    """קובץ CSV/XLSX → רשימת שורות טקסט (לקבצים שלא נטענו ל-datastore)."""
    import io
    rows = []
    if fmt == 'CSV':
        import csv
        for enc in ('utf-8-sig', 'cp1255', 'utf-8'):
            try:
                rows = list(csv.reader(io.StringIO(content.decode(enc))))
                break
            except Exception:
                continue
    elif fmt == 'XLS':
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            for ws in wb.sheets():
                for i in range(ws.nrows):
                    rows.append([str(c.value).strip() for c in ws.row(i)])
        except Exception as e:
            print('  פענוח xls נכשל:', e)
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for ws in wb.worksheets:
                for r in ws.iter_rows(values_only=True):
                    rows.append(['' if c is None else str(c).strip() for c in r])
        except Exception as e:
            print('  פענוח xlsx נכשל:', e)
    return rows


def harvest_rows(rows, by_city):
    """מאתר שורת כותרת עם עמודת אשכול ועמודת שם, וקוטף את הערכים.
    בקובצי הלמ"ס הכותרת לפעמים מפוצלת לשתי שורות — מנסים גם צירוף זוגות."""
    added = 0
    ci = ni = None

    def find_cols(row):
        cc = [j for j, c in enumerate(row) if 'אשכול' in c and 'דירוג' not in c]
        cn = [j for j, c in enumerate(row)
              if any(w in c for w in ('שם הרשות', 'שם רשות', 'שם היישוב', 'שם יישוב', 'שם הישוב', 'רשות מקומית'))
              or c.strip() == 'שם']
        return cc, cn

    for i, row in enumerate(rows):
        cand_c, cand_n = find_cols(row)
        if cand_c and cand_n:
            ci, ni, start = cand_c[0], cand_n[0], i + 1
            break
        if cand_c and i + 1 < len(rows):     # כותרת דו-שורתית
            _, cn2 = find_cols(rows[i + 1])
            if cn2:
                ci, ni, start = cand_c[0], cn2[0], i + 2
                break
    if ci is None:
        return 0
    for row in rows[start:]:
        if len(row) <= max(ci, ni):
            continue
        name = str(row[ni]).strip()
        try:
            c = int(float(str(row[ci]).strip()))
        except (TypeError, ValueError):
            continue
        if name and 1 <= c <= 10:
            if name not in by_city:
                added += 1
            by_city.setdefault(name, {'c': c})
    return added


def main():
    pkgs = []
    for q in ('מדד חברתי-כלכלי', 'מדד חברתי כלכלי', 'socio-economic',
              'אשכול חברתי', 'אשכול כלכלי רשויות', 'רשויות מקומיות', 'נתוני רשויות',
              'אשכול', 'עיריות', 'אשכול כלכלי'):
        try:
            res = call('package_search', q=q, rows=32).get('results', [])
            print(f'חיפוש "{q}": {len(res)}')
            pkgs += res
        except Exception as e:
            print('חיפוש נכשל:', q, e)
    # מאגר העיריות הוא "אח" של מאגר המועצות — אצל אותו מפרסם; סורקים את כל
    # המאגרים של כל ארגון שפרסם מאגר "אשכול" כלשהו, ואת ארגון הלמ"ס עצמו
    orgs = {(p.get('organization') or {}).get('name')
            for p in pkgs if 'אשכול' in (p.get('title') or '')}
    try:
        for o in call('organization_list', all_fields='true', limit=400):
            hay = (o.get('name', '') + ' ' + o.get('display_name', '') + ' ' + o.get('title', '')).lower()
            if any(w in hay for w in ('cbs', 'statist', 'סטטיסטיקה', 'למ"ס')):
                orgs.add(o['name'])
    except Exception as e:
        print('רשימת ארגונים נכשלה:', e)
    for o in sorted(o for o in orgs if o):
        try:
            more = call('package_search', fq=f'organization:{o}', rows=100).get('results', [])
            print(f'ארגון {o}: {len(more)} מאגרים')
            pkgs += more
        except Exception as e:
            print('סריקת ארגון נכשלה:', o, e)
    seen = set()
    cands = []
    for p in pkgs:
        if p['id'] in seen:
            continue
        seen.add(p['id'])
        title = p.get('title', '')
        # האשכול של הערים חי במאגרי "נתוני רשויות מקומיות" שאין בכותרתם
        # "חברתי-כלכלי" — הסינון האמיתי הוא קיום שדה אשכול במשאב עצמו
        rel = (('חברתי' in title and 'כלכלי' in title) or 'אשכול' in title
               or 'עיריות' in title
               or 'רשויות מקומיות' in title or 'ישובים' in title or 'יישובים' in title)
        if not rel:
            continue
        ym = re.findall(r'(20\d\d)', title + ' ' + (p.get('notes') or '')[:200])
        year = max(map(int, ym)) if ym else 0
        cands.append((year, title, p))
    cands.sort(key=lambda x: -x[0])
    cands = cands[:60]
    print('מועמדים:', [(y, t[:60]) for y, t, _ in cands[:25]])

    # ממזגים את כל המשאבים המתאימים: הערים (רשויות מקומיות) והיישובים
    # שבתוך מועצות אזוריות יושבים במשאבים/מאגרים נפרדים
    by_city = {}
    srcs = []
    best_year = 0
    for year, title, p in cands:
        for res in p.get('resources', []):
            if not res.get('datastore_active'):
                continue
            rid = res['id']
            try:
                probe = call('datastore_search', resource_id=rid, limit=5)
            except Exception as e:
                print('  דילוג על משאב:', rid, e)
                continue
            fields = probe.get('fields', [])
            # חלק מהמאגרים עם שדות בעברית וחלק באנגלית (ESHKOL / LOCALITY)
            f_cluster = find_field(fields, 'אשכול') or find_field(fields, 'ESHKOL')
            f_name = (find_field(fields, 'HEBREW', 'LOCALITY')
                      or find_field(fields, 'HEBREW', 'MUNICIP')
                      or find_field(fields, 'HEBREW', 'NAME')
                      or find_field(fields, 'שם', 'ישוב') or find_field(fields, 'שם', 'יישוב')
                      or find_field(fields, 'שם', 'רשות') or find_field(fields, 'שם'))
            f_rank = find_field(fields, 'דירוג') or find_field(fields, 'RANK')
            if not (f_cluster and f_name):
                print('  אין שדות מתאימים:', title[:45], [f.get('id') for f in fields][:9])
                continue
            rows = []
            offset = 0
            try:
                while True:
                    chunk = call('datastore_search', resource_id=rid, limit=5000, offset=offset)
                    rec = chunk.get('records', [])
                    rows += rec
                    if len(rec) < 5000:
                        break
                    offset += 5000
            except Exception as e:
                print('  משיכת שורות נקטעה:', title[:40], e)
                continue
            added = 0
            for r in rows:
                name = str(r.get(f_name) or '').strip()
                try:
                    c = int(float(r.get(f_cluster)))
                except (TypeError, ValueError):
                    continue
                if not name or not (1 <= c <= 10):
                    continue
                ent = {'c': c}
                if f_rank:
                    try:
                        ent['r'] = int(float(r.get(f_rank)))
                    except (TypeError, ValueError):
                        pass
                if name not in by_city:
                    added += 1
                by_city.setdefault(name, ent)
            if added:
                print(f'מוזג: {title[:60]} | {res.get("name","")[:40]} | +{added} (שדות {f_name}/{f_cluster})')
                srcs.append(title[:70])
                best_year = max(best_year, year)
    # מסלול ב': קבצים שמצורפים למאגרים אך לא נטענו ל-datastore (שם מסתתר
    # בדרך כלל קובץ העיריות) — מורידים ומפענחים ידנית
    big_missing = not any(c in by_city for c in ('תל אביב - יפו', 'תל אביב-יפו', 'ירושלים'))
    if big_missing:
        for year, title, p in cands:
            if 'אשכול' not in title and not ('חברתי' in title and 'כלכלי' in title):
                continue
            for res in p.get('resources', []):
                fmt = (res.get('format') or '').upper()
                if fmt not in ('CSV', 'XLSX', 'XLS') or res.get('datastore_active'):
                    continue
                rname = res.get('name') or res.get('url', '').rsplit('/', 1)[-1]
                try:
                    content = fetch_bytes(res['url'])
                except Exception as e:
                    print('  הורדה נכשלה:', rname[:50], e)
                    continue
                added = harvest_rows(parse_tabular(content, fmt), by_city)
                print(f'  קובץ {rname[:60]} ({fmt}): +{added}')
                if added:
                    srcs.append(f'{title[:60]} ({rname[:40]})')
                    best_year = max(best_year, year)

    # מסלול ג': אתר הלמ"ס עצמו — ב-data.gov.il אין את אשכולות העיריות בכלל
    # (נבדק: גם ארגון lamas וגם חיפושים ישירים). קובצי הפרסום הרשמי של
    # "אפיון רשויות מקומיות לפי הרמה החברתית-כלכלית" יושבים ב-doclib.
    big_missing = not any(c in by_city for c in ('תל אביב - יפו', 'תל אביב-יפו', 'ירושלים'))
    if big_missing:
        # נתיבים מאומתים (החיפוש איתר את ה-PDF-ים המקבילים בדיוק בנתיבים אלה):
        # הודעה 230/2024 = המדד לשנת 2021; פרסום socio_eco19_1903 = מדד 2019
        cbs_files = [
            ('https://www.cbs.gov.il/he/mediarelease/DocLib/2024/230/24_24_230t1.xlsx', 2021),
            ('https://www.cbs.gov.il/he/mediarelease/DocLib/2024/230/24_24_230t2.xlsx', 2021),
            ('https://www.cbs.gov.il/he/publications/doclib/2023/socio_eco19_1903/t01.xlsx', 2019),
            ('https://www.cbs.gov.il/he/publications/doclib/2023/socio_eco19_1903/t02.xlsx', 2019),
        ]
        for page in ('https://www.cbs.gov.il/he/subjects/Pages/%D7%9E%D7%93%D7%93-%D7%97%D7%91%D7%A8%D7%AA%D7%99-%D7%9B%D7%9C%D7%9B%D7%9C%D7%99-%D7%A9%D7%9C-%D7%94%D7%A8%D7%A9%D7%95%D7%99%D7%95%D7%AA-%D7%94%D7%9E%D7%A7%D7%95%D7%9E%D7%99%D7%95%D7%AA.aspx',):
            try:
                html = fetch_bytes(page).decode('utf-8', 'ignore')
                hrefs = re.findall(r'href="([^"]+)"', html)
                xl = [h for h in hrefs if re.search(r'\.xlsx?($|\?)', h, re.I)]
                print(f'עמוד למ"ס: {len(hrefs)} קישורים, {len(xl)} אקסל')
                for h in xl[:15]:
                    u = h if h.startswith('http') else 'https://www.cbs.gov.il' + h
                    if u not in [c[0] for c in cbs_files]:
                        yr = 2021 if '2021' in u or '/230/' in u else 2019
                        cbs_files.append((u, yr))
                        print('  קישור מהעמוד:', u[-70:])
            except Exception as e:
                print('עמוד למ"ס נכשל:', page[:60], e)
        for u, yr in cbs_files[:14]:
            fname = u.rsplit('/', 1)[-1]
            fmt = 'XLS' if u.lower().endswith('.xls') else 'XLSX'
            try:
                content = fetch_bytes(u)
            except Exception as e:
                print('  למ"ס הורדה נכשלה:', fname, e)
                continue
            if not content.startswith(b'PK') and fmt == 'XLSX':
                print(f'  למ"ס {fname}: לא אקסל ({content[:60]!r})')
                continue
            added = harvest_rows(parse_tabular(content, fmt), by_city)
            print(f'  למ"ס {fname}: +{added}')
            if added:
                srcs.append(f'למ"ס — המדד החברתי-כלכלי {yr} ({fname})')
                best_year = max(best_year, yr)
            if any(c in by_city for c in ('תל אביב - יפו', 'תל אביב-יפו', 'ירושלים')):
                break

    for probe_city in ('תל אביב - יפו', 'תל אביב-יפו', 'ירושלים', 'דימונה', 'קרית גת', 'קריית גת'):
        if probe_city in by_city:
            print('בדיקה:', probe_city, '→ אשכול', by_city[probe_city]['c'])
    if len(by_city) < 200:
        sys.exit(f'מעט מדי יישובים ({len(by_city)}) — לא שומר')
    out = {'year': best_year or None, 'source': ' + '.join(dict.fromkeys(srcs)),
           'n': len(by_city), 'by_city': by_city}
    with open('parks/data/socio.json', 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False)
    print(f'נשמרו {len(by_city)} יישובים ורשויות (שנת {best_year})')


if __name__ == '__main__':
    main()
