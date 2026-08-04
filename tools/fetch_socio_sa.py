#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""אשכול חברתי-כלכלי לפי אזור סטטיסטי — לכל אזור תעשייה.

הנתון העירוני מטעה לפעמים (אזור תעשייה יכול לשבת באזור סטטיסטי חזק בעיר
חלשה, ולהפך — קרית מלאכי אשכול 3, אך האזור הסטטיסטי של אזור התעשייה
שלה אשכול 6). המקורות:
  1. לוח האזורים הסטטיסטיים של פרסום המדד 2021 (למ"ס, הודעה 230/2024)
  2. שכבת הגבולות של האזורים הסטטיסטיים מעמוד השכבות של הלמ"ס (GDB,
     מומר עם ogr2ogr ל-GeoJSON ב-WGS84)
לכל אזור תעשייה: נקודת המרכז ← האזור הסטטיסטי המכיל ← האשכול.
פלט: parks/data/socio-sa.json {"zones": {pNNN.json: {"c":…, "code":…}}}
רץ על runner של GitHub (דורש gdal-bin; לסביבה המקומית אין רשת לדומיינים).
"""
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.parse
import urllib.request

UA = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

# שכבת "מדד חברתי כלכלי 2021 לאזורים סטטיסטיים 2011" — אותה שכבה שמוצגת
# ב-govmap — מפורסמת גם כפריט ArcGIS ציבורי עם האשכול בתוך המאפיינים
ARCGIS_ITEM = '5814e892a6494b3488f9bccf67e36687'


def fetch(url, timeout=300):
    # כתובות עם עברית (שם השכבה ב-ArcGIS) חייבות קידוד אחוזים
    url = urllib.parse.quote(url, safe=':/?&=%')
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def load_sa_from_arcgis():
    """השכבה כ-GeoJSON ישירות משירות ה-FeatureServer — בלי GDB בכלל."""
    meta = json.loads(fetch(f'https://www.arcgis.com/sharing/rest/content/items/{ARCGIS_ITEM}?f=json'))
    url = meta.get('url')
    print('פריט ArcGIS:', meta.get('title', '')[:60], '| סוג:', meta.get('type'), '| שירות:', url)
    if not url:
        # פריט מסוג Web Map — כתובות השכבות בתוך נתוני הפריט (operationalLayers)
        data = json.loads(fetch(f'https://www.arcgis.com/sharing/rest/content/items/{ARCGIS_ITEM}/data?f=json'))
        for ly in data.get('operationalLayers', []):
            u = ly.get('url') or ''
            print('  שכבה במפה:', ly.get('title', '')[:50], '→', u[:100])
            if 'Server' in u:
                url = u
                break
        if not url:
            return None
    if not re.search(r'/\d+$', url):
        url += '/0'
    url = url.rstrip('/')
    # מפרידים את אינדקס השכבה מהשירות לצורך שאילתות
    print('שכבת המקור:', url)
    feats = []
    offset = 0
    while True:
        q = (f'{url}/query?where=1%3D1&outFields=*&outSR=4326&f=geojson'
             f'&resultOffset={offset}&resultRecordCount=1000')
        d = json.loads(fetch(q))
        if 'error' in d:
            print('שגיאת שירות:', str(d["error"])[:150])
            return None
        fs = d.get('features', [])
        feats += fs
        if len(fs) < 1000:
            break
        offset += 1000
    print(f'{len(feats)} אזורים סטטיסטיים מהשירות')
    return feats or None


# ---------- 1. לוח האשכולות לפי אזור סטטיסטי ----------
def load_sa_clusters():
    from openpyxl import load_workbook
    import io
    for tno in ('t3', 't4', 't2'):
        url = f'https://www.cbs.gov.il/he/mediarelease/DocLib/2024/230/24_24_230{tno}.xlsx'
        try:
            content = fetch(url)
        except Exception as e:
            print(f'  {tno}: הורדה נכשלה — {e}')
            continue
        if not content.startswith(b'PK'):
            print(f'  {tno}: לא אקסל')
            continue
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                rows.append(['' if c is None else str(c).strip() for c in r])
        # שורת הכותרת: יש בה גם "אשכול" וגם "סטטיסטי"
        hdr_i = ci = None
        for i, row in enumerate(rows):
            has_c = [j for j, c in enumerate(row) if 'אשכול' in c]
            has_sa = any('סטטיסטי' in c for c in row)
            if has_c and has_sa:
                hdr_i, ci = i, has_c[0]
                break
        if hdr_i is None:
            print(f'  {tno}: אין לוח אזורים סטטיסטיים')
            continue
        hdr = rows[hdr_i]
        # קוד משולב (סמל יישוב+אזור, 8 ספרות) או שני שדות נפרדים
        c_comb = next((j for j, c in enumerate(hdr) if 'סטטיסטי' in c and 'סמל' in c), None)
        c_sym = next((j for j, c in enumerate(hdr) if 'סמל' in c and 'סטטיסטי' not in c), None)
        c_sa = next((j for j, c in enumerate(hdr) if 'סטטיסטי' in c and 'סמל' not in c), None)
        print(f'  {tno}: כותרת בשורה {hdr_i}: {[h[:20] for h in hdr if h][:8]}')
        out = {}
        for row in rows[hdr_i + 1:]:
            try:
                cl = int(float(row[ci]))
                if not 1 <= cl <= 10:
                    continue
                if c_comb is not None and row[c_comb]:
                    code = int(float(row[c_comb]))
                elif c_sym is not None and c_sa is not None:
                    code = int(float(row[c_sym])) * 10000 + int(float(row[c_sa]))
                else:
                    continue
                out[code] = cl
            except (ValueError, TypeError, IndexError):
                continue
        if len(out) > 500:
            print(f'  {tno}: {len(out)} אזורים סטטיסטיים עם אשכול')
            return out
        print(f'  {tno}: רק {len(out)} שורות — ממשיך לנסות')
    return None


# ---------- 2. שכבת הגבולות ----------
def load_sa_polygons(tmp):
    page = fetch('https://www.cbs.gov.il/he/Pages/geo-layers.aspx').decode('utf-8', 'replace')
    cands = [h for h in re.findall(r'href="([^"]+\.zip)"', page, re.I)
             if 'statistical' in h.lower() or 'סטטיסטי' in urllib.parse.unquote(h)]
    print('שכבות מועמדות:', [c[-60:] for c in cands])
    if not cands:
        return None
    url = cands[0] if cands[0].startswith('http') else 'https://www.cbs.gov.il' + cands[0]
    zp = os.path.join(tmp, 'sa.zip')
    open(zp, 'wb').write(fetch(url, timeout=600))
    print(f'שכבה הורדה: {os.path.getsize(zp)/1e6:.0f}MB')
    import zipfile
    ex = os.path.join(tmp, 'sa')
    zipfile.ZipFile(zp).extractall(ex)
    # מאתרים gdb או shp
    src = None
    for root, dirs, files in os.walk(ex):
        for d in dirs:
            if d.lower().endswith('.gdb'):
                src = os.path.join(root, d)
        for fl in files:
            if fl.lower().endswith('.shp') and src is None:
                src = os.path.join(root, fl)
    if not src:
        print('אין gdb/shp בארכיון')
        return None
    print('מקור:', src)
    gj = os.path.join(tmp, 'sa.json')
    r = subprocess.run(['ogr2ogr', '-f', 'GeoJSON', '-t_srs', 'EPSG:4326', gj, src],
                       capture_output=True, text=True)
    if r.returncode:
        print('ogr2ogr נכשל:', r.stderr[:400])
        return None
    return json.load(open(gj, encoding='utf-8'))


def ring_contains(ring, x, y):
    inside = False
    n = len(ring)
    j = n - 1
    for i in range(n):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if (yi > y) != (yj > y) and x < (xj - xi) * (y - yi) / (yj - yi) + xi:
            inside = not inside
        j = i
    return inside


def poly_contains(geom, x, y):
    polys = geom['coordinates'] if geom['type'] == 'MultiPolygon' else [geom['coordinates']]
    for poly in polys:
        if not poly:
            continue
        if ring_contains(poly[0], x, y):
            if all(not ring_contains(h, x, y) for h in poly[1:]):
                return True
    return False


def find_code_key(p0):
    for k, v in p0.items():
        if 'STAT' in k.upper():
            try:
                if int(float(v)) > 100000:
                    return k
            except (TypeError, ValueError):
                continue
    return None


def build_items(feats, cluster_of):
    """[(bbox, geometry, cluster, code)] עם סינון מה שאין לו אשכול."""
    items = []
    for f in feats:
        g = f.get('geometry')
        if not g or g.get('type') not in ('Polygon', 'MultiPolygon'):
            continue
        cl, code = cluster_of(f.get('properties', {}))
        if not cl:
            continue
        xs, ys = [], []
        polys = g['coordinates'] if g['type'] == 'MultiPolygon' else [g['coordinates']]
        for poly in polys:
            for pt in poly[0]:
                xs.append(pt[0]); ys.append(pt[1])
        if xs:
            items.append((min(xs), min(ys), max(xs), max(ys), g, cl, code))
    return items


def main():
    items = None
    # מסלול א': שירות ה-ArcGIS — האשכול כבר במאפיינים
    try:
        feats = load_sa_from_arcgis()
    except Exception as e:
        print('ArcGIS נכשל:', e)
        feats = None
    if feats:
        p0 = feats[0].get('properties', {})
        print('שדות השכבה:', list(p0.keys())[:15])
        ckey = next((k for k in p0 if 'אשכול' in k or any(w in k.lower() for w in ('eshkol', 'cluster', 'madad'))), None)
        if ckey is None:   # זיהוי לפי טווח הערכים 1–10
            for k in p0:
                vals = [f['properties'].get(k) for f in feats[:300]]
                good = [v for v in vals if isinstance(v, (int, float)) and float(v).is_integer() and 1 <= v <= 10]
                if len(good) > 200:
                    ckey = k
                    break
        code_key = find_code_key(p0)
        print('שדה האשכול:', ckey, '| שדה הקוד:', code_key)
        if ckey:
            def cluster_of(props):
                try:
                    cl = int(float(props.get(ckey)))
                except (TypeError, ValueError):
                    return None, None
                code = None
                if code_key:
                    try:
                        code = int(float(props.get(code_key)))
                    except (TypeError, ValueError):
                        pass
                return (cl if 1 <= cl <= 10 else None), code
            items = build_items(feats, cluster_of)
            print(f'{len(items)} אזורים עם אשכול מהשירות')

    # מסלול ב' (נפילה): לוח אקסל + שכבת GDB מאתר הלמ"ס
    if not items:
        clusters = load_sa_clusters()
        if not clusters:
            sys.exit('אין לוח אשכולות לאזורים סטטיסטיים — לא ממשיך')
        with tempfile.TemporaryDirectory() as tmp:
            gj = load_sa_polygons(tmp)
            if not gj:
                sys.exit('אין שכבת גבולות')
            feats = gj.get('features', [])
            key = find_code_key(feats[0].get('properties', {}))
            if key is None:
                sys.exit(f'לא זוהה שדה קוד: {list(feats[0].get("properties", {}).keys())[:12]}')

            def cluster_of(props):
                try:
                    code = int(float(props.get(key)))
                except (TypeError, ValueError):
                    return None, None
                return clusters.get(code), code
            items = build_items(feats, cluster_of)

    parks = json.load(open('parks/data/parks.json', encoding='utf-8'))
    zones = {}
    for p in parks:
        x, y = p['lo'], p['la']
        for x0, y0, x1, y1, g, cl, code in items:
            if x0 <= x <= x1 and y0 <= y <= y1 and poly_contains(g, x, y):
                zones[p['f']] = {'c': cl}
                if code:
                    zones[p['f']]['code'] = code
                break
    print(f'{len(zones)} מתוך {len(parks)} אזורי תעשייה שויכו לאזור סטטיסטי עם אשכול')
    if len(zones) < 50:
        sys.exit('מעט מדי שיוכים — לא שומר')
    json.dump({'year': 2021, 'source': 'למ"ס — המדד החברתי-כלכלי 2021 לאזורים סטטיסטיים (שכבת ה-GIS הרשמית)', 'n': len(zones), 'zones': zones},
              open('parks/data/socio-sa.json', 'w', encoding='utf-8'), ensure_ascii=False)
    print('נשמר parks/data/socio-sa.json')


if __name__ == '__main__':
    main()
